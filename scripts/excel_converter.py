from openpyxl import Workbook, load_workbook
import sys

#f_name = '   '

#print('>>>STUNDENPLAN-KONVERTER<<<\n\nBitte legen sie die Tabelle und die Programmdatei in den selben Ordern!\n\n')

#print('Bitte geben sie den Namen der Stpl.-datei an: ')
#f_name = input()

#print('Bitte geben sie den Namen der Ausgabedatei an: ')
#f_out_name = input()

alphabet = ['B','C','D','E','F']

f_name = sys.argv[1]
f_out_name = sys.argv[2]

wb_in = load_workbook(filename = str(f_name))
page = wb_in['seite1']
wb_out = Workbook()
dest_filename = str(f_out_name)
ws1_out = wb_out.active
ws1_out.title = 'tabelle1'
ws1_out['A1'] = 'Name'
ws1_out['B1'] = 'Klasse'
ws1_out['C1'] = 'Lehrer'


def read(column, row):
    first = str(page[column+str(row)].value)
    sec = first.split('\n')
    out = list()
    for i in sec:
        out.append(i.split(';'))
    return out

zeile = 2

for letters in alphabet:
    for zellen in range(2,12):
        liste = read(letters, zellen)
        for s in liste:
            print('Zeile ' + str(zeile-1)+ ' ' + str(s))
            ws1_out['A'+str(zeile)] = str(s[0])
            ws1_out['B'+str(zeile)] = str(s[1])
            ws1_out['C'+str(zeile)] = str(s[2])
            wb_out.save(filename = dest_filename)
            zeile = zeile + 1


#print('Die Tabelle wurde abgespeichert als ' + dest_filename)





